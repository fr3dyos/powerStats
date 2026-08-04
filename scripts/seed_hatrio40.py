"""Seed the HatRio 40° 2019 tournament into the Supabase database.

Idempotent: skips insertion if a tournament named "HatRio 40° 2019" already
exists. Uses the same SQLAlchemy session factory as the FastAPI backend
(``database.py`` + ``DATABASE_URL`` env var).

Run from the repo root with the backend venv:

    .\\.venv\\Scripts\\python.exe scripts\\seed_hatrio40.py

The script expects ``.env`` at the repo root (loaded by ``database.py``).
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

# Allow running from repo root: add the repo dir to sys.path so the
# ``database`` / ``models`` modules resolve.
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from database import SessionLocal  # noqa: E402
import models  # noqa: E402


EXCEL_PATH = REPO_ROOT / "Power Stats - HatRio40° 2019 - Estatisticas.xlsx"

TOURNAMENT_NAME = "HatRio 40° 2019"
TOURNAMENT_START = datetime(2019, 6, 20, tzinfo=timezone.utc)
TOURNAMENT_END = datetime(2019, 6, 23, 23, 59, tzinfo=timezone.utc)
TOURNAMENT_LOCATION = "Rio de Janeiro"
TOURNAMENT_DESCRIPTION = (
    "HatRio 40° 2019 — seeded from the source statistics spreadsheet. "
    "Round-robin + bracket, 6 color teams."
)

TEAM_NAMES = ["Amarelo", "Azul", "Cinza", "Rosa", "Verde", "Vermelho"]

RR_HOUR_TO_MIN = {
    "9h": 9 * 60,
    "10h": 10 * 60,
    "11h": 11 * 60,
    "12h": 12 * 60,
    "13h": 13 * 60,
    "14h": 14 * 60,
    "15h": 15 * 60,
    "16h": 16 * 60,
    "11h30": 11 * 60 + 30,
    "12h30": 12 * 60 + 30,
    "13h30": 13 * 60 + 30,
    "14h30": 14 * 60 + 30,
    "15h30": 15 * 60 + 30,
    "16h30": 16 * 60 + 30,
}


def _int(value, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _str(value) -> str:
    return "" if value is None else str(value).strip()


def _day_offset(day_value, hour_label: str, fallback_idx: int) -> int:
    if isinstance(day_value, datetime):
        return (day_value.date() - TOURNAMENT_START.date()).days
    # Fallback: position-based. RR has 6 games on day 0, 6 on day 1,
    # 3 on day 2. Finais adds 7 more across days 2/3.
    if fallback_idx <= 6:
        return 0
    if fallback_idx <= 12:
        return 1
    if fallback_idx <= 15:
        return 2
    return 3


def game_start_time(day_offset: int, hour_label: str) -> datetime:
    minute_of_day = RR_HOUR_TO_MIN.get(hour_label, 12 * 60)
    base_date = TOURNAMENT_START.date() + timedelta(days=day_offset)
    return datetime(
        base_date.year,
        base_date.month,
        base_date.day,
        minute_of_day // 60,
        minute_of_day % 60,
        tzinfo=timezone.utc,
    )


def parse_rr_games(rows: list[tuple]) -> list[dict]:
    """Yield 15 round-robin games from the RR sheet's data rows.

    Data rows occupy indices 9..26 in the spreadsheet, with sparse
    rows for "Almoço" lunch and blank gutters. We skip rows with no
    ``Jogo`` number and tolerate the rest.
    """
    games: list[dict] = []
    for idx, row in enumerate(rows[9:27], start=1):
        number = _int(row[4])
        home = _str(row[5])
        home_score = _int(row[6])
        away_score = _int(row[7])
        away = _str(row[8])
        if not (home and away and number):
            continue
        games.append(
            {
                "number": number,
                "day_offset": _day_offset(row[2], row[3], idx),
                "hour_label": _str(row[3]),
                "home": home,
                "away": away,
                "home_score": home_score,
                "away_score": away_score,
            }
        )
    return games


def parse_finais_games(rows: list[tuple]) -> list[dict]:
    """Yield 7 bracket games from the Finais sheet's data rows.

    Layout (data row): ``Fase | Dia | Hora | Jogo | I time | PONTOS | II time``.
    Column offsets: ``jogo=5, home=6, home_score=7, away_score=8, away=9``.
    Games 16..22 occupy indices 8, 9, 11, 12, 13, 14, 15 (row 10 is a
    bracket placeholder row that we skip).
    """
    games: list[dict] = []
    indices = [8, 9, 11, 12, 13, 14, 15]
    for idx, row_idx in enumerate(indices, start=16):
        row = rows[row_idx]
        day = row[3]
        hour = _str(row[4])
        number = _int(row[5])
        home = _str(row[6])
        home_score = _int(row[7])
        away_score = _int(row[8])
        away = _str(row[9])
        if not (home and away and number):
            continue
        games.append(
            {
                "number": number,
                "day_offset": _day_offset(day, hour, idx),
                "hour_label": hour,
                "home": home,
                "away": away,
                "home_score": home_score,
                "away_score": away_score,
            }
        )
    return games


def parse_compiled_players(rows: list[tuple]) -> list[dict]:
    """Yield player rows from the Compiled sheet."""
    players: list[dict] = []
    # Row 1 = header; data starts at row 2.
    for row in rows[2:]:
        team = _str(row[0])
        if not team or team not in TEAM_NAMES:
            continue
        jersey = _int(row[1])
        full_name = _str(row[2])
        if not full_name:
            continue
        parts = full_name.split(maxsplit=1)
        first = parts[0]
        last = parts[1] if len(parts) > 1 else ""
        players.append(
            {
                "team": team,
                "jersey": jersey,
                "first_name": first,
                "last_name": last,
                "assists": _int(row[4]),
                "goals": _int(row[5]),
                "defenses": _int(row[6]),
            }
        )
    return players


def parse_log_events(rows: list[tuple]) -> list[dict]:
    """Yield per-game events from the Log compiled sheet.

    Header at row index 2 (0-indexed) reads:
    ``(None, None, 'Team A', 'Team B', 'Jogo ', 'Time', 'Action',
      'Score Team A', 'Score B', 'Assist/Defence', 'Scorer', ...)``.
    Data rows start at index 3.
    """
    pattern = re.compile(r"^(\d+)/(\d+)/(\d+) @ (\d+):(\d+):(\d+)$")
    events: list[dict] = []
    for row in rows[3:]:
        if not row or len(row) < 11:
            continue
        jogo = _int(row[4])
        ts_raw = _str(row[5])
        action = _str(row[6])
        if not action:
            continue
        m = pattern.match(ts_raw)
        if m is None:
            continue
        d, mo, y, hh, mm, ss = (int(x) for x in m.groups())
        ts = datetime(y, mo, d, hh, mm, ss, tzinfo=timezone.utc)
        events.append(
            {
                "jogo": jogo,
                "ts": ts,
                "action": action,
                "scorer": _str(row[10]),
                "assist": _str(row[9]),
                "home_team": _str(row[2]),
                "away_team": _str(row[3]),
            }
        )
    return events


def capitalize_color(name: str) -> str:
    return name[:1].upper() + name[1:].lower()


def find_player(
    players_by_key: dict[tuple[str, str], models.Player],
    team: models.Team,
    name: str,
) -> models.Player | None:
    if not name:
        return None
    parts = name.split(maxsplit=1)
    first = parts[0]
    last = parts[1] if len(parts) > 1 else ""
    return players_by_key.get((team.name, f"{first} {last}"))


def first_half_cutoff(events: list[dict]) -> datetime | None:
    if not events:
        return None
    times = sorted(e["ts"] for e in events)
    return times[len(times) // 2]


def seed() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rr_rows = list(workbook["RR"].iter_rows(values_only=True))
    fin_rows = list(workbook["Finais"].iter_rows(values_only=True))
    compiled_rows = list(workbook["Compiled"].iter_rows(values_only=True))
    log_rows = list(workbook["Log compiled"].iter_rows(values_only=True))

    rr_games = parse_rr_games(rr_rows)
    fin_games = parse_finais_games(fin_rows)
    parsed_players = parse_compiled_players(compiled_rows)
    log_events = parse_log_events(log_rows)

    print(
        f"Parsed {len(rr_games)} RR games, {len(fin_games)} bracket games, "
        f"{len(parsed_players)} players, {len(log_events)} log events."
    )

    session = SessionLocal()
    try:
        existing = (
            session.query(models.Tournament)
            .filter(models.Tournament.name == TOURNAMENT_NAME)
            .first()
        )
        if existing is not None:
            print(
                f"Tournament '{TOURNAMENT_NAME}' already exists "
                f"(id={existing.id}). Skipping seed."
            )
            return

        # 1) Tournament
        tournament = models.Tournament(
            name=TOURNAMENT_NAME,
            start_date=TOURNAMENT_START,
            end_date=TOURNAMENT_END,
            location=TOURNAMENT_LOCATION,
            description=TOURNAMENT_DESCRIPTION,
        )
        session.add(tournament)
        session.flush()

        # 2) Teams
        teams_by_name: dict[str, models.Team] = {}
        for name in TEAM_NAMES:
            team = models.Team(name=name, tournament_id=tournament.id)
            session.add(team)
            teams_by_name[name] = team
        session.flush()

        # 3) Players (build a key→player map for log lookups later)
        players_by_id: dict[int, models.Player] = {}
        players_by_key: dict[tuple[str, str], models.Player] = {}
        for p in parsed_players:
            team = teams_by_name[p["team"]]
            player = models.Player(
                first_name=p["first_name"],
                last_name=p["last_name"],
                jersey_number=p["jersey"] or None,
                team_id=team.id,
            )
            session.add(player)
            session.flush()
            players_by_id[player.id] = player
            players_by_key[(p["team"], f"{p['first_name']} {p['last_name']}")] = player

        # 4) Games — insert directly via Core so we never load a Game ORM
        # object (the model declares columns like is_placement that don't
        # exist in the live DB; the ORM mapper would try to flush them
        # back to the table and fail).
        # Build a sanitized Table view that only contains the legacy columns
        # actually present in the live DB, so SQLAlchemy never references
        # the newer ones (is_placement, phase_id, group_id, etc.).
        from sqlalchemy import Column as _Col, MetaData as _MetaData, Table as _Table  # noqa: E402

        live_cols = {
            "id",
            "tournament_id",
            "home_team_id",
            "away_team_id",
            "start_time",
            "end_time",
            "home_score",
            "away_score",
            "game_rule",
            "time_limit",
            "score_limit",
            "field_number",
            "is_completed",
            "created_at",
            "updated_at",
        }
        legacy_table = _Table(
            "games",
            _MetaData(),
            *[
                c.copy()
                for c in models.Game.__table__.columns
                if c.name in live_cols
            ],
        )
        games: dict[int, int] = {}  # number -> game_id
        for spec in rr_games + fin_games:
            start_time = game_start_time(spec["day_offset"], spec["hour_label"])
            result = session.execute(
                legacy_table.insert().values(
                    tournament_id=tournament.id,
                    home_team_id=teams_by_name[spec["home"]].id,
                    away_team_id=teams_by_name[spec["away"]].id,
                    start_time=start_time,
                    end_time=start_time + timedelta(minutes=30),
                    home_score=spec["home_score"],
                    away_score=spec["away_score"],
                    game_rule=models.GameRuleEnum.TIME_LIMIT.value,
                    time_limit=30,
                    field_number=1,
                    is_completed=True,
                )
            )
            games[spec["number"]] = result.inserted_primary_key[0]
        session.flush()

        # 5) Events from Log compiled
        per_game_events: dict[int, list[dict]] = defaultdict(list)
        for e in log_events:
            per_game_events[e["jogo"]].append(e)

        for jogo, events in per_game_events.items():
            game_id = games.get(jogo)
            if game_id is None:
                continue
            # Look up the game start time so we can record elapsed seconds.
            spec = next(
                (g for g in rr_games + fin_games if g["number"] == jogo),
                None,
            )
            start_time = (
                game_start_time(spec["day_offset"], spec["hour_label"])
                if spec
                else None
            )

            def elapsed(ts, _start=start_time):
                if _start is None:
                    return None
                return int((ts - _start).total_seconds())

            events_sorted = sorted(events, key=lambda x: x["ts"])
            cutoff = first_half_cutoff(events_sorted)
            for ev in events_sorted:
                period = 1 if cutoff is None or ev["ts"] <= cutoff else 2
                action_lower = ev["action"].lower()

                if action_lower.startswith("goal "):
                    team_name = action_lower.replace("goal ", "").strip()
                    team = teams_by_name.get(capitalize_color(team_name))
                    if team is None:
                        continue
                    scorer = find_player(
                        players_by_key, team, ev["scorer"]
                    )
                    if scorer is None:
                        continue
                    session.add(
                        models.GameEvent(
                            game_id=game_id,
                            player_id=scorer.id,
                            event_type=models.GameEventTypeEnum.GOAL,
                            points=1,
                            time_elapsed=elapsed(ev["ts"]),
                            period=period,
                        )
                    )
                    if ev["assist"]:
                        assist_player = find_player(
                            players_by_key, team, ev["assist"]
                        )
                        if assist_player is not None:
                            session.add(
                                models.GameEvent(
                                    game_id=game_id,
                                    player_id=assist_player.id,
                                    event_type=models.GameEventTypeEnum.ASSIST,
                                    points=1,
                                    time_elapsed=elapsed(ev["ts"]),
                                    period=period,
                                )
                            )

                elif action_lower.startswith("defence "):
                    team_name = action_lower.replace("defence ", "").strip()
                    team = teams_by_name.get(capitalize_color(team_name))
                    if team is None:
                        continue
                    # The log puts the defender in the "Assist/Defence" column.
                    defender = find_player(
                        players_by_key, team, ev["assist"]
                    )
                    if defender is None:
                        continue
                    session.add(
                        models.GameEvent(
                            game_id=game_id,
                            player_id=defender.id,
                            event_type=models.GameEventTypeEnum.DEFENSE,
                            points=1,
                            time_elapsed=elapsed(ev["ts"]),
                            period=period,
                        )
                    )

        # 6) Per-tournament stats from the Compiled totals.
        games_per_team: dict[int, int] = defaultdict(int)
        for spec in rr_games + fin_games:
            home_id = teams_by_name[spec["home"]].id
            away_id = teams_by_name[spec["away"]].id
            games_per_team[home_id] += 1
            games_per_team[away_id] += 1

        # Index parsed Compiled rows by (team, "First Last") so we can pull
        # the goals/assists/defenses totals for each inserted player.
        totals_by_key: dict[tuple[str, str], dict[str, int]] = {}
        for p in parsed_players:
            key = (p["team"], f"{p['first_name']} {p['last_name']}")
            totals_by_key[key] = {
                "goals": p["goals"],
                "assists": p["assists"],
                "defenses": p["defenses"],
            }

        for player in players_by_id.values():
            team = next(t for t in teams_by_name.values() if t.id == player.team_id)
            totals = totals_by_key.get(
                (team.name, f"{player.first_name} {player.last_name}"),
                {"goals": 0, "assists": 0, "defenses": 0},
            )
            session.add(
                models.PlayerTournamentStats(
                    player_id=player.id,
                    tournament_id=tournament.id,
                    games_played=games_per_team.get(player.team_id, 0),
                    goals=totals["goals"],
                    assists=totals["assists"],
                    defenses=totals["defenses"],
                    goals_conceded=0,
                )
            )

        session.commit()
        print(
            f"Seeded tournament id={tournament.id} with "
            f"{len(teams_by_name)} teams, {len(players_by_id)} players, "
            f"{len(games)} games."
        )
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


if __name__ == "__main__":
    seed()